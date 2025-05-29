import pandas as pd
import openpyxl
import os

file_path = "/home/ubuntu/upload/app.xlsx"
analysis_output_file = "/home/ubuntu/excel_analysis.txt"

analysis_results = []

try:
    # Use pandas to get sheet names and basic data structure
    xls = pd.ExcelFile(file_path)
    sheet_names = xls.sheet_names
    analysis_results.append(f"Análise da Planilha: {os.path.basename(file_path)}\n")
    analysis_results.append(f"Total de Abas: {len(sheet_names)}\n")
    analysis_results.append(f"Nomes das Abas: {', '.join(sheet_names)}\n")
    analysis_results.append("-" * 30 + "\n")

    # Use openpyxl for more detailed analysis (formulas, charts, VBA)
    # keep_vba=True to check for macros, data_only=False to read formulas
    workbook = openpyxl.load_workbook(file_path, read_only=False, keep_vba=True, data_only=False)

    analysis_results.append("Análise Detalhada por Aba:\n")

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        analysis_results.append(f"\n--- Aba: {sheet_name} ---\n")
        analysis_results.append(f"Dimensões (linhas x colunas): {sheet.max_row} x {sheet.max_column}\n")
        analysis_results.append(f"Intervalo de Células: {sheet.dimensions}\n")

        # Check for tables
        if hasattr(sheet, 'tables') and sheet.tables:
            analysis_results.append(f"Tabelas Estruturadas Encontradas: {len(sheet.tables)}\n")
            for table_name, table_ref in sheet.tables.items():
                analysis_results.append(f"  - Nome: {table_name}, Intervalo: {table_ref}\n")
        else:
            analysis_results.append("Tabelas Estruturadas Encontradas: Nenhuma\n")

        # Check for formulas (sample a few cells)
        formula_count = 0
        formula_examples = []
        # Iterate through a sample of rows/cols to find formulas without loading everything
        # Limit check for performance, adjust as needed
        max_rows_to_check = min(sheet.max_row, 100) if sheet.max_row else 0
        max_cols_to_check = min(sheet.max_column, 50) if sheet.max_column else 0
        if max_rows_to_check > 0 and max_cols_to_check > 0:
            for row in range(1, max_rows_to_check + 1):
                 for col in range(1, max_cols_to_check + 1):
                    cell = sheet.cell(row=row, column=col)
                    # Check if cell exists and has a formula
                    if cell and cell.data_type == 'f':
                        formula_count += 1
                        if len(formula_examples) < 5: # Store first 5 examples
                             formula_value = cell.value
                             # Ensure formula value is a string before slicing
                             if isinstance(formula_value, str):
                                 # Correctly format the f-string for examples
                                 formula_examples.append(f"   - Célula {cell.coordinate}: ={formula_value[:100]}{'...' if len(formula_value) > 100 else ''}")
                             else:
                                 # Handle cases where formula might not be string (though unlikely for 'f' type)
                                 formula_examples.append(f"   - Célula {cell.coordinate}: [Não foi possível exibir a fórmula]")

        analysis_results.append(f"Fórmulas Encontradas (amostra): {'Sim (%d encontradas na amostra)' % formula_count if formula_count > 0 else 'Não'}\n")
        if formula_examples:
             # Correctly append the header for formula examples
             analysis_results.append("  Exemplos de Fórmulas (início):\n")
             # Extend with each example followed by a newline
             analysis_results.extend([f"{example}\n" for example in formula_examples])

        # Check for charts
        chart_count = 0
        if hasattr(sheet, '_charts'):
             chart_count = len(sheet._charts)
        analysis_results.append(f"Gráficos Encontrados: {chart_count}\n")
        # Note: Getting chart details (type, data source) with openpyxl is complex.

        # Check for images
        image_count = 0
        if hasattr(sheet, '_images'):
            image_count = len(sheet._images)
        analysis_results.append(f"Imagens Encontradas: {image_count}\n")

        # Check for Pivot Tables
        pivot_count = 0
        if hasattr(sheet, 'pivot_tables') and sheet.pivot_tables:
             pivot_count = len(sheet.pivot_tables)
        analysis_results.append(f"Tabelas Dinâmicas Encontradas: {pivot_count}\n")

    # Check for VBA Macros
    has_vba = workbook.vba_archive is not None
    analysis_results.append("-" * 30 + "\n")
    analysis_results.append(f"Contém Macros VBA: {'Sim' if has_vba else 'Não'}\n")
    if has_vba:
        analysis_results.append("  (Nota: A análise detalhada do código VBA requer ferramentas específicas e não é feita aqui.)\n")

    # Write analysis to file
    with open(analysis_output_file, "w", encoding="utf-8") as f:
        f.writelines(analysis_results)

    print(f"Análise concluída. Resultados salvos em {analysis_output_file}")

except FileNotFoundError:
    print(f"Erro: Arquivo não encontrado em {file_path}")
except ImportError as e:
    print(f"Erro de importação: {e}. Verifique se pandas e openpyxl estão instalados.")
except Exception as e:
    # Provide more specific error feedback if possible
    if 'zipfile.BadZipFile' in str(e):
        print(f"Erro: O arquivo '{os.path.basename(file_path)}' pode estar corrompido ou não é um arquivo Excel válido (.xlsx).")
    else:
        print(f"Ocorreu um erro inesperado durante a análise: {e}")

